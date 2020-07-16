# built-in
import sys
import os
import time
import re
from datetime import datetime as dtm
if sys.version_info[0] == 3:
    import tkinter as tk                
    from tkinter import messagebox as tkMessageBox
    from tkinter import font  as tkfont 
    from tkinter import ttk             
    from tkinter import filedialog
else:
    import Tkinter as tk     
    import tkFont as tkfont  
    import ttk
    import tkFileDialog as filedialog
    import tkMessageBox
    
# external
import openpyxl
import plotly.graph_objects as go
import plotly


# time.strftime("%Y%m%d%H%M%S")

def parseLogLines(lines):
    stats = []
    header = ["Time"]
    data = {}
    time_regex = "(\d+)\/(\d+)\/(\d+)\s+(\d+)\:(\d+)\:(\d+)"
    curr_time = ""

    # going through all lines in log
    for line in lines:
        result = re.search(time_regex, line)
        # current line is a timestamp
        if result:
            ctg = result.groups()
            curr_time = dtm(int(ctg[2]), int(ctg[1]), int(ctg[0]), int(ctg[3]), int(ctg[4]), int(ctg[5]))
            if data:
                #print(stats[-1])
                if curr_time != data["Time"]:
                    data = {"Time":curr_time, "Comments":""}
            else:
                data = {"Time":curr_time, "Comments":""}
        elif "AVISO 1" in line or "AVISO SISTEMA NORMALIZADO gc" in line and curr_time:
            stats[-1]["Comments"] = line.replace("[MONITOR] ", "")
        elif "AVISO" in line and curr_time:
            data["Comments"] = line.replace("[MONITOR] ", "")
        # current line has a value to be stored
        elif ":" in line and curr_time:
            l = line.replace("[MONITOR] ", "").split(":")
            k = l[0].strip()
            v = l[1].strip()
            if k not in header:
                header.append(k)
            if "-" in v:
                data[k] = float(v.split("-")[0])
            else:
                data[k] = float(v.split("%")[0])
            if k == "- Quantidade HTTP":
                stats.append(data)
                '''if "Heap %" in data:
                    stats.append(data)
                else:
                    stats[-1].update(data)'''
        # current line has a value to be stored
        elif "%" in line and curr_time:
            if "heap" in k:
                if "Heap %" not in header:
                    header.append("Heap %")
                data["Heap %"] = float(line.replace("[MONITOR] ", "").split("%")[0])
            elif "perm" in k:
                if "Perm %" not in header:
                    header.append("Perm %")
                data["Perm %"] = float(line.replace("[MONITOR] ", "").split("%")[0])
            elif "Garbage" in k:
                if "Garbage %" not in header:
                    header.append("Garbage %")
                data["Garbage %"] = float(line.replace("[MONITOR] ", "").split("%")[0])
    header.append("Comments")
    return stats, header

if __name__ == "__main__":
    print("Select a valid log file")
    print("\n\n")
    tk.Tk().withdraw()
    file_path = filedialog.askopenfile()
    f = open(file_path.name, "r")
    lines = f.readlines()
    print("Parsing logs...")
    print("\n\n")
    stats, header = parseLogLines(lines)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    c = ws['B2']
    ws.freeze_panes = c
    
    
    # chr(ord('A')+0) = 'A'
    h_width = [20, 15, 15, 10, 15, 15, 10, 15, 15, 10, 15, 15, 50]
    for idx,h in enumerate(header):
        if idx >= len(h_width):
            ws.column_dimensions[chr(ord('A')+idx)].width = 15
        else:
            ws.column_dimensions[chr(ord('A')+idx)].width = h_width[idx]    
        #ws[chr(ord('A')+idx)+"1"] = h.replace(" -", "")
        ws.cell(row=1, column=1+idx).value = h.replace("-", "").strip()
        ws.auto_filter.ref = ws.dimensions
        
    t = []    
    heap_perc = []
    perm_perc = []
    garb_perc = []
    try:
        for i,stat in enumerate(stats):
            for j,h in enumerate(header):
                if h in stat:
                    ws.cell(row=2+i, column=1+j).value = stat[h]
                    if "Time" in h:
                        t.append(stat[h])
                    elif "Heap %" in h:
                        heap_perc.append(stat[h])
                    elif "Perm %" in h:
                        perm_perc.append(stat[h])
                    elif "Garbage %" in h:
                        garb_perc.append(stat[h])
                else:
                    ws.cell(row=2+i, column=1+j).value = ""
    except:
        print("Events exceeded limit... stopping")
    report_name = os.path.basename(file_path.name).split(".")[0] + "_" + time.strftime("%Y%m%d%H%M%S")
    wb.save("./Reports/" + report_name + ".xlsx")
    

    #fig = go.Figure(data=[go.Scatter(x=t, y=heap_perc)])
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=t, y=heap_perc, name="Heap %"))
    fig.add_trace(go.Scatter(x=t, y=perm_perc, name="Perm %"))
    fig.add_trace(go.Scatter(x=t, y=garb_perc, name="Garbage %"))
    plotly.offline.plot(fig, filename = "./Reports/" + report_name + ".html", auto_open=False)
    print("\n\n")
    if sys.version_info[0] == 3:
        a = input("Finished! Press ENTER to close")
    else:
        a = raw_input("Finished! Press ENTER to close")